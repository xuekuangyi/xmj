import os

import openpyxl
from django.db import IntegrityError
from django.http import HttpResponse
from openpyxl import load_workbook

from core.models import *
import datetime


# 文件导入功能

# `models.GoodsCost.objects.filter(soNo=800290872520005250).values('soNo')[1]['soNo']`

# 获取成本的方法
def inputCommand():
	bizType = input("请选择导入数据类型，编码表如下\n"
	                "1:导入供商数据；\n2:预留；\n"
	                "3:导入门店商品数据；\n4:组合商品导入；\n5:导入商品供货关系;\n"
	                "10:导入采购单；\n"
	                "12：导入收货单抬头；\n13:收货单明细；\n"
	                "15:导入销售订单抬头；\n16:导入销售单明细\n")
	
	path = input("请输入导入文件的路径")
	sheetNo = input("请输入要导入的sheet")
	importData(bizType, path, sheetNo)
	return HttpResponse("上传数据成功")
	command = input('继续y 退出n')
	return command


# 读取excel方法读取调用时指定文件及sheet号，返回sheet_obj对象——一个可迭代的对象
def readExcel(path, sheetNo):
	# 获取文件对象
	file_path = os.path.join(path)  # 只能用绝对路径，如何解决？
	# Workbook就是一个excel文档对象
	wb_obj = load_workbook(file_path)
	print("文件读取成功")
	# 获取sheet对象
	sheet_obj = wb_obj.worksheets[int(sheetNo)]  # 再定位到某个sheet，并实例化一个sheet对象
	return sheet_obj


def createExcelFile(sheetName):
	wb_obj = openpyxl.Workbook()
	sheet = wb_obj.active
	sheet.title = f"{sheetName}"
	


def importData(bizType, path, sheetNo):

	# 获取到要被上传的sheet对象
	sheet_obj = readExcel(path, sheetNo)
	print(sheet_obj)
	count = 0
	# 对该对象进行逐行循环遍历，赋值给对应的数据库表；
	for row in sheet_obj.iter_rows(min_row=2):
		
		count += 1

		# 导入供商数据
		if bizType == '1':
			try:
				Supplier.objects.create(
					supplierNo=int(row[0].value),
					supplierName=row[1].value,
					deliveryDays=row[5].value
				)
			except IntegrityError as error:
				print(f'错误类型{error},{row[0]}已存在')
		# 判断业务类型为2时，导入
		elif bizType == '2':
			pass
		

		# 导入门店商品数据，门店商品数据会向Goods和GoodsSpec对象中拆分
		elif bizType == '3':
			# print("正常进入类型3")
			# 将数据导入到商品表
			
				try:
					Goods.objects.create(
						spuId=row[2].value,
						goodsName=row[4].value
					)
					
					# 注意区分querySet和实例
					# print(Goods.objects.filter(spuId=row[2].value).first())
				except IntegrityError as error:
					pass
					# print(f'重复SPU{row[2]}')
			
				GoodsSpec.objects.create(
					goods=Goods.objects.filter(spuId=row[2].value).first(),
					skuId=row[3].value,
					upc=row[20].value,
					specName=row[7].value
				)
				try:
					StoreGoods.objects.create(
						store=Store.objects.filter(storeNo=1031328).first(),
						# 只负责第一次处理
						sku=GoodsSpec.objects.filter(skuId=row[3].value).first(),
						retailPrice=int(float(row[14].value) * 1000),
						status=0,
						infoCreateTime=0,
						infoUpdateTime=0
					
					)
				except ValueError as error:
					if row[14].value == '':
						print(f'{row[14]}售价未配置')
						StoreGoods.objects.create(
							store=Store.objects.filter(storeNo=1031328).first(),
							# 只负责第一次处理
							sku=GoodsSpec.objects.filter(skuId=row[3].value).first(),
							retailPrice=0,
							status=0,
							infoCreateTime=0,
							infoUpdateTime=0
						
						)
		
		# 组合商品导入
		elif bizType == '4':
			try:
				Bom.objects.create(
				storeGoods=StoreGoods.objects.filter(sku=row[3].value).first(),
				unitGoods=StoreGoods.objects.filter(sku=row[9].value).first(),
				usage=float(row[14].value)

				)
			
			except IntegrityError as error:
				pass
				# print(row[3].value)
				# print(row[9].value)
			
		# 导入商品供货关系；
		elif bizType == '5':
			PurchaseRelations.objects.create(
				relationNo=int(row[0].value),
				store=Store.objects.filter(storeNo=row[1].value).first(),
				supplier=Supplier.objects.filter(supplierNo=row[3].value).first(),
				arrivalDays=row[5].value,
				sku=GoodsSpec.objects.filter(skuId=row[8].value).first(),
				purUnit=row[10].value,
				isDefault=row[14].value,
				purPrice=int(float(row[15].value)*100)
			)
		
		# 导入采购单
		elif bizType == '10':
			pass
		
		
		# 导入收货单抬头
		elif bizType == '12':
			try:
				ReceiptBill.objects.create(
					receiptBillNo=row[0].value,
					store = Store.objects.filter(storeName=row[1].value).first(),
					supplier = Supplier.objects.filter(supplierName=row[2].value).first(),
					source = row[3].value,
					sourceBillNo = row[4].value,
					stauts=row[5].value,
					creator=row[10].value
				)
			except IntegrityError as error:
				print(f'找不到供商或“发货方”不为供商:{row[2].value}')
		
		# 导入收货单明细
		elif bizType == '13':
			# 只导入收货单状态为“已完成”的收货单的明细
			try:
				if ReceiptBill.objects.filter(receiptBillNo=row[0].value).exists() and ReceiptBill.objects.filter(receiptBillNo=row[0].value).first().stauts == '已完成':
				# if row[3].value == '采购单':
					ReceiptBillItems.objects.create(
						receiptBill=ReceiptBill.objects.filter(receiptBillNo=row[0].value).first(),
						sku=GoodsSpec.objects.filter(skuId=row[4].value).first(),
						unit=row[8].value,
						sendQty=row[9].value,
						netRecQty=row[11].value,
						recQty=row[12].value,
						grandTotalRefund=row[13].value,
						offset=row[14].value,
						actPurchasePrice=PurchaseRelations.objects.filter(sku=GoodsSpec.objects.filter(skuId=row[4].value).first()).first().purPrice,
						purchasePriceByBill=int(float(row[15].value)*100),
						purchaseAmount=int(float(row[16].value)*100),
						netPurchaseAmount=int(float(row[18].value)*100),
						receiver=row[20].value,
						receiveTime=row[21].value
					)
				else:
					print(f'{row[0].value}未完成或未开始收货')
			except ValueError as error01:
				print(f'数据异常ValueError【{error01}】,单据号:{row[0].value},发货方:{row[2].value}，单据类型{row[3].value}')
			except IntegrityError as error02:
				# print(f'数据异常IntegrityError,【{error02}】|单据号:{row[0].value}|,发货方:{row[2].value}|，单据类型{row[3].value}|,净收货量{row[11].value}')
				print(f'{row[0].value}已关单且本行商品{row[5].value}无净收货|,发货方:{row[2].value}|，单据类型{row[3].value}|,净收货量{row[11].value}')
			except AttributeError as error03:
				print(f'数据异常AttributeError,【{error03}】|单据号:{row[0].value},发货方{row[2].value}|，单据类型{row[3].value}|,净收货量{row[11].value}')

		# 导入销售订单抬头
		elif bizType == '15':
			# 转义订单状态
			if row[21] == "已完成":
				transfer = 3
			# 导入销售单抬头：
			SellOrder.objects.create(
				
				store=Store.objects.filter(storeName=row[2].value).first(),
				soNo=row[0].value,
				soSn=int(row[1].value),
				
				totalAmount=int(float(row[8].value)*100),
				totalGoodsAmount=int(float(row[9].value)*100),
				realPayTotal=int(float(row[10].value)*100),
				realMerTotal=int(float(row[11].value)*100),
				deliveryFee=int(float(row[12].value)*100),
				lunchboxFee=int(float(row[13].value)*100),
			
				platformServiceFee=int(float(row[14].value)*100),
				merActFee=int(float(row[15].value)*100),
				platformActFee=int(float(row[16].value)*100),

				status=row[21].value,
				deliveryStatus=row[22].value,
				
				createTime=datetime.datetime.strptime(row[25].value, "%Y-%m-%d %H:%M:%S"),
				finishTime=row[27].value,
				refundTime=row[28].value,
			)


		# 导入销售单明细
		elif bizType == '16':
			# 导入销售商品明细
			try:
				if row[29].value != '-':
					gpt = int(float(row[29].value)*100)
				else:
					gpt = 0
					
				if row[30].value != '-':
					gpm = int(float(row[29].value)*100)
				else:
					gpm = 0
					
				if row[31].value != '-':
					gpp = int(float(row[29].value)*100)
				else:
					gpp = 0
					
				
				
				SoDetails.objects.create(
					refSo=SellOrder.objects.filter(soNo=row[2].value).first(),
					
					sku=GoodsSpec.objects.filter(skuId=row[18].value).first(),
					upc=row[19].value,
					specName=row[21].value,
					unitName=row[23].value,
					
					qty=float(row[25].value),
					isRefundGoods=row[32].value,
					refundQty=float(row[33].value),
					netQty=float(row[25].value)-float(row[33].value),
	
					sellGoodsPrice=int(float(row[26].value)*100),
					# refPurRela=PurchaseRelations.objects.filter(
					# 	store=Store.objects.filter(storeName=row[2].value).first(),
					# 	# sku=row[18].value
					# ).first(),
					# # 需要将first()替换为找价格最小值
					# # refGoodsCost=PurchaseRelations.objects.filter(
					# # 	store=row[2].value,
					# # 	sku=row[18].value
					# # ).first(),
					
					goodsSellAmount=int(float(row[27].value)*100),
					goodsRealPayAmount=int(float(row[28].value)*100),
					goodsRefundAmount=int(float(row[37].value)*100),
					
					goodsPerkTotalAmount=gpt,
					goodsPerkMerAmount=gpm,
					goodsPerkPlatformAmount=gpp,
					
					refundTime=row[38].value
	
				)
			except ValueError as error:
				print(f'跳过异常{error}|订单号：{row[2].value},sku:{row[18].value}')
				
			except IntegrityError as error:
				print(f'{error}|订单号：{row[2].value},sku:{row[18].value}')
			
			except AttributeError as error:
				print(f'异常类型【{error}】订单号:{row[2].value},sku:{row[18].value}')
		
		elif bizType == '17':
			# 预处理
			file_status = 0
			if row[3].value == '出勤':
				file_status = 0
			elif row[3].value == '缺勤':
				file_status = 1
			
			
			OnSale.objects.create(
				date_by_day=row[0].value,
				sku = GoodsSpec.objects.filter(skuId=row[18].value).first(),
				yesterday_show_hours=
				status=
			
			
			)
			
			
	print(f'共执行[{count}]次')
			
			
